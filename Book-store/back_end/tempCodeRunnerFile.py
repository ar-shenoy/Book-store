from flask import Flask, request, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask_cors import CORS
import random
import string
import pandas as pd
import csv
from flask_caching import Cache  # Import Flask-Caching
import joblib  # Import joblib for loading the model
import numpy as np  # Import numpy for data manipulation
from model import predict_fraud  # Import the fraud detection function from model.py

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Initialize Flask-Caching
cache = Cache(app, config={'CACHE_TYPE': 'simple'})  # Cache configuration

# Load or create the users Excel file
def load_users():
    try:
        workbook = openpyxl.load_workbook('resources/users.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Email', 'Username', 'Password'])  # Column headers
        workbook.save('resources/users.xlsx')
    return workbook, sheet

# Load books data from the CSV file
def load_books():
    books = []
    try:
        with open('resources/books.csv', mode='r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                row['Price'] = row['Price'].replace('Price Starting at ', '').replace('$', '').strip()
                row['Description'] = row.get('Description', 'No description available')
                books.append(row)
        return books
    except FileNotFoundError:
        print("books.csv file not found.")
        return []
    except Exception as e:
        print(f"Error loading books: {e}")
        return []

# Send reset password email
def send_reset_email(email, reset_link):
    sender_email = "your_email@example.com"  # Replace with your email
    sender_password = "your_password"  # Replace with your email password
    recipient_email = email

    # Create the email content
    subject = "Password Reset Request"
    body = f"Click on the following link to reset your password: {reset_link}"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    # Connect to the SMTP server and send the email
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)  # Use your SMTP server
        server.starttls()
        server.login(sender_email, sender_password)
        text = msg.as_string()
        server.sendmail(sender_email, recipient_email, text)
        server.quit()
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

# Generate a random password reset token
def generate_reset_token():
    return ''.join(random.choices(string.ascii_letters + string.digits, k=20))

# Fraud detection route (credit card validation)
@app.route('/check_credit_card', methods=['POST'])
def check_credit_card():
    data = request.get_json()
    card_features = data.get('features', [])

    if not card_features:
        return jsonify({"message": "No credit card features provided."}), 400

    # Call the fraud detection model to predict
    result = predict_fraud(card_features)
    
    # Return the result to the frontend
    if 'error' in result:
        return jsonify({"message": result["error"]}), 500
    
    fraud_status = "fraud" if result["fraud"] else "valid"
    return jsonify({
        "message": f"The credit card is {fraud_status}.",
        "fraud_probability": result["probability"]
    }), 200

# Signup route
@app.route('/signup', methods=['POST'])
def signup():
    data = request.get_json()
    email = data['email']
    username = data['username']
    password = data['password']
    password2 = data['password2']

    if password != password2:
        return jsonify({"message": "Passwords do not match"}), 400

    # Check if user already exists
    workbook, sheet = load_users()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == username or row[0] == email:
            return jsonify({"message": "User already exists"}), 400

    # Hash the password before saving
    hashed_password = generate_password_hash(password)

    # Add user to Excel file
    sheet.append([email, username, hashed_password])
    workbook.save('resources/users.xlsx')

    return jsonify({"message": "User registered successfully"}), 200

# Login route
@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data['username']
    password = data['password']

    workbook, sheet = load_users()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            if check_password_hash(row[2], password):
                return jsonify({"message": "Login successful"}), 200
            else:
                return jsonify({"message": "Invalid credentials"}), 400
    return jsonify({"message": "User not found"}), 400

# Forgot Password route
@app.route('/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json()
    email = data['email']

    workbook, sheet = load_users()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == email:
            reset_token = generate_reset_token()
            reset_link = f"http://localhost:5000/reset-password?token={reset_token}"
            
            # Store the token (In a real-world app, you'd store this in a database with expiration time)
            # For simplicity, we're not saving it here
            if send_reset_email(email, reset_link):
                return jsonify({"message": "Password reset link sent to your email"}), 200

    return jsonify({"message": "Email not found"}), 400

# Reset Password route (with token)
@app.route('/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json()
    token = data['token']
    new_password = data['new_password']
    
    # Logic to validate the token (In a real-world app, check the token from the database)
    if token:  # For simplicity, we're skipping token validation here
        # Update the password (In a real-world app, you'd search for the user by email and update)
        # Hash the new password
        hashed_password = generate_password_hash(new_password)
        # You'd update the password in the database or file here
        return jsonify({"message": "Password reset successfully"}), 200
    else:
        return jsonify({"message": "Invalid or expired token"}), 400

# Get Books route without caching for pagination
@app.route('/get_books', methods=['GET'])
def get_books():
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))

    books = load_books()

    # Ensure valid page and per_page values
    if page < 1:
        page = 1
    if per_page < 1:
        per_page = 20

    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    print(f"Fetching books from {start_index} to {end_index}")  # Log the range

    # Paginate the books
    paginated_books = books[start_index:end_index]

    return jsonify(paginated_books)

@app.route('/search_books', methods=['GET'])
def search_books():
    search_query = request.args.get('query', '').lower()
    books = load_books()
    filtered_books = [book for book in books if search_query in book['Title'].lower() or search_query in book['Authors'].lower()]
    return jsonify(filtered_books)

# Subscribe route
@app.route('/subscribe', methods=['POST'])
def subscribe():
    data = request.get_json()
    email = data['email']

    # Logic to handle subscription (e.g., save to a file or database)
    # For now, we simply return a success message
    return jsonify({"message": f"Subscribed successfully with {email}"}), 200

# Contact Us route
@app.route('/contact-us', methods=['POST'])
def contact_us():
    data = request.get_json()
    name = data['name']
    email = data['email']
    message = data['message']

    # Logic to handle contact us (e.g., save the message or send an email)
    # For now, we simply return a success message
    return jsonify({"message": f"Message from {name} received. We will get back to you shortly!"}), 200

# Start the Flask app
if __name__ == '__main__':
    app.run(debug=True)
